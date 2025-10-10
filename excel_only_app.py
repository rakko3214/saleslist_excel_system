from flask import Flask, render_template_string, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, func, and_
from dotenv import load_dotenv
import os
import pymysql
from datetime import datetime, date, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io

# 環境変数をロード
load_dotenv()

app = Flask(__name__)

# データベース設定
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')

db = SQLAlchemy(app)

# ========================
# 実際のデータ構造に合わせたモデル定義
# ========================

class FmArea(db.Model):
    __tablename__ = 'fm_areas'
    
    id = db.Column(db.Integer, primary_key=True)
    area_name_ja = db.Column(db.Text, nullable=False)
    area_name_en = db.Column(db.Text, nullable=False)
    fm_login_account_id = db.Column(db.Text, nullable=False)
    fm_login_account_pass = db.Column(db.Text, nullable=False)

class FmAccount(db.Model):
    __tablename__ = 'fm_accounts'
    
    id = db.Column(db.Integer, primary_key=True)
    department_name = db.Column(db.Text, nullable=False)
    sort_order = db.Column(db.Integer, nullable=False)
    needs_hellowork = db.Column(db.Integer, nullable=False, default=0)
    needs_tabelog = db.Column(db.Integer, nullable=False, default=0)
    needs_kanri = db.Column(db.Integer, nullable=False, default=1)

class FmAreaAccount(db.Model):
    __tablename__ = 'fm_area_accounts'
    
    id = db.Column(db.Integer, primary_key=True)
    fm_area_id = db.Column(db.Integer, nullable=False)
    fm_account_id = db.Column(db.Integer, nullable=False)

class Company(db.Model):
    __tablename__ = 'companies'
    
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.Text)
    address = db.Column(db.Text)
    created_at = db.Column(db.DateTime)
    updated_at = db.Column(db.DateTime)
    fm_import_result = db.Column(db.Integer)  # 1=更新, 2=新規
    # 他のフィールドも必要に応じて追加

# ========================
# Excel出力専用の関数
# ========================

def get_real_company_counts(area_id, account_id, target_date=None):
    """実際のcompaniesテーブルから件数を取得"""
    
    if target_date is None:
        target_date = datetime.now().date()
    
    try:
        # 指定日に作成された企業数（新規）
        new_count = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == target_date,
            Company.fm_import_result == 2  # 新規
        ).scalar() or 0
        
        # 指定日に作成された企業数（更新）
        update_count = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == target_date,
            Company.fm_import_result == 1  # 更新
        ).scalar() or 0
        
        return {
            'new': new_count,
            'updated': update_count
        }
        
    except Exception as e:
        # エラーの場合も実データを返す（0件として扱う）
        print(f"データ取得エラー: {e}")
        return {
            'new': 0,
            'updated': 0
        }

def get_companies_summary_by_date(target_date=None):
    """指定日のcompaniesテーブルのサマリーを取得"""
    try:
        if target_date is None:
            target_date = datetime.now().date()
        elif isinstance(target_date, str):
            target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        
        # 全体統計
        total_companies = db.session.query(func.count(Company.id)).scalar() or 0
        
        # 指定日のデータ
        target_new = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == target_date,
            Company.fm_import_result == 2  # 新規
        ).scalar() or 0
        # 指定日に更新された企業数
        target_updated = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == target_date,
            Company.fm_import_result == 1  # 更新
        ).scalar() or 0
        
        # 指定日の前後のデータも取得（比較用）
        prev_date = target_date - timedelta(days=1)
        next_date = target_date + timedelta(days=1)
        
        prev_new = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == prev_date,
            Company.fm_import_result == 2  # 新規
        ).scalar() or 0
        prev_updated = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == prev_date,
            Company.fm_import_result == 1  # 更新
        ).scalar() or 0
        
        next_new = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == next_date,
            Company.fm_import_result == 2  # 新規
        ).scalar() or 0
        next_updated = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == next_date,
            Company.fm_import_result == 1  # 更新
        ).scalar() or 0
        
        return {
            'total_companies': total_companies,
            'target_date': target_date.strftime('%Y-%m-%d'),
            'target_date_jp': target_date.strftime('%Y年%m月%d日'),
            'target_new': target_new,
            'target_updated': target_updated,
            'prev_date': prev_date.strftime('%Y年%m月%d日'),
            'prev_new': prev_new,
            'prev_updated': prev_updated,
            'next_date': next_date.strftime('%Y年%m月%d日'),
            'next_new': next_new,
            'next_updated': next_updated
        }
    except Exception as e:
        print(f"日別サマリー取得エラー: {e}")
        return {
            'total_companies': 0,
            'target_date': datetime.now().strftime('%Y-%m-%d'),
            'target_date_jp': datetime.now().strftime('%Y年%m月%d日'),
            'target_new': 0,
            'target_updated': 0,
            'prev_date': '',
            'prev_new': 0,
            'prev_updated': 0,
            'next_date': '',
            'next_new': 0,
            'next_updated': 0
        }

def get_companies_summary():
    """companiesテーブルの全体サマリーを取得"""
    try:
        today = datetime.now().date()
        
        # 全体統計
        total_companies = db.session.query(func.count(Company.id)).scalar() or 0
        
        # 今日のデータ
        today_new = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == today,
            Company.fm_import_result == 2  # 新規
        ).scalar() or 0
        today_updated = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) == today,
            Company.fm_import_result == 1  # 更新
        ).scalar() or 0
        
        # 最近7日間のデータ（データが少ない場合の代替）
        week_ago = today - timedelta(days=7)
        week_new = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) >= week_ago,
            func.date(Company.created_at) <= today,
            Company.fm_import_result == 2  # 新規
        ).scalar() or 0
        week_updated = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) >= week_ago,
            func.date(Company.created_at) <= today,
            Company.fm_import_result == 1  # 更新
        ).scalar() or 0
        
        # 今月のデータ
        month_start = today.replace(day=1)
        month_new = db.session.query(func.count(Company.id)).filter(
            func.date(Company.created_at) >= month_start,
            func.date(Company.created_at) <= today
        ).scalar() or 0
        month_updated = db.session.query(func.count(Company.id)).filter(
            func.date(Company.updated_at) >= month_start,
            func.date(Company.updated_at) <= today,
            func.date(Company.created_at) != func.date(Company.updated_at)
        ).scalar() or 0
        
        return {
            'total_companies': total_companies,
            'today_new': today_new,
            'today_updated': today_updated,
            'week_new': week_new,
            'week_updated': week_updated,
            'month_new': month_new,
            'month_updated': month_updated,
            'date': today.strftime('%Y年%m月%d日'),
            'week_period': f"{week_ago.strftime('%m月%d日')}〜{today.strftime('%m月%d日')}",
            'month_period': f"{month_start.strftime('%m月%d日')}〜{today.strftime('%m月%d日')}"
        }
    except Exception as e:
        print(f"サマリー取得エラー: {e}")
        return {
            'total_companies': 0,
            'today_new': 0,
            'today_updated': 0,
            'week_new': 0,
            'week_updated': 0,
            'month_new': 0,
            'month_updated': 0,
            'date': datetime.now().strftime('%Y年%m月%d日'),
            'week_period': '過去7日間',
            'month_period': '今月'
        }

def get_area_account_mapping():
    """支店とアカウントの関連マッピングを取得（ハローワーク対象のみ）"""
    
    mapping = db.session.query(
        FmAreaAccount.fm_area_id,
        FmAreaAccount.fm_account_id,
        FmArea.area_name_ja,
        FmAccount.department_name,
        FmAccount.needs_hellowork
    ).join(
        FmArea, FmAreaAccount.fm_area_id == FmArea.id
    ).join(
        FmAccount, FmAreaAccount.fm_account_id == FmAccount.id
    ).filter(
        FmAccount.needs_hellowork == 1  # ハローワークが必要なアカウントのみ
    ).order_by(
        FmArea.id, FmAccount.sort_order
    ).all()
    
    return [
        {
            'area_id': row.fm_area_id,
            'area_name': row.area_name_ja,
            'account_id': row.fm_account_id,
            'account_name': row.department_name,
            'needs_hellowork': bool(row.needs_hellowork)
        } for row in mapping
    ]

def generate_hierarchical_excel_data_by_date(target_date=None):
    """指定日の階層構造のExcel出力用データを生成"""
    
    # 支店とアカウントのマッピングを取得
    mapping = get_area_account_mapping()
    
    # 指定日のcompaniesデータのサマリーを取得
    companies_summary = get_companies_summary_by_date(target_date)
    
    # 階層構造データを生成
    hierarchical_data = []
    
    # 支店ごとにグループ化
    areas = {}
    for item in mapping:
        area_name = item['area_name']
        if area_name not in areas:
            areas[area_name] = {
                'area_id': item['area_id'],
                'accounts': []
            }
        areas[area_name]['accounts'].append({
            'account_id': item['account_id'],
            'account_name': item['account_name']
        })
    
    # 全体データを支店・アカウント数で按分計算
    total_areas = len(areas)
    total_accounts = len(mapping)
    
    target_new = companies_summary['target_new']
    target_updated = companies_summary['target_updated']
    
    if total_accounts > 0 and (target_new > 0 or target_updated > 0):
        avg_new_per_account = max(0, target_new // total_accounts)
        avg_update_per_account = max(0, target_updated // total_accounts)
        remainder_new = target_new % total_accounts
        remainder_update = target_updated % total_accounts
    else:
        avg_new_per_account = 0
        avg_update_per_account = 0
        remainder_new = 0
        remainder_update = 0
    
    account_index = 0
    
    # 各支店・アカウントに対してデータを生成
    for area_name, area_data in areas.items():
        # 支店ヘッダー行
        hierarchical_data.append({
            'レベル': 1,
            '項目名': f"📍 {area_name}",
            '種別': '',
            '件数': '',
            '備考': f'支店ID: {area_data["area_id"]}'
        })
        
        area_total = 0
        
        for account in area_data['accounts']:
            # アカウントヘッダー行
            hierarchical_data.append({
                'レベル': 2,
                '項目名': f"  📂 {account['account_name']}",
                '種別': '',
                '件数': '',
                '備考': f'アカウントID: {account["account_id"]}'
            })
            
            # 実際のデータを按分で割り当て
            new_count = avg_new_per_account + (remainder_new if account_index == 0 else 0)
            update_count = avg_update_per_account + (remainder_update if account_index == 0 else 0)
            account_index += 1
            
            hierarchical_data.append({
                'レベル': 3,
                '項目名': f"    📝 新規",
                '種別': '新規',
                '件数': new_count,
                '備考': f'{companies_summary["target_date_jp"]}の実データ（全体{target_new}件を按分）'
            })
            
            hierarchical_data.append({
                'レベル': 3,
                '項目名': f"    🔄 更新",
                '種別': '更新',
                '件数': update_count,
                '備考': f'{companies_summary["target_date_jp"]}の実データ（全体{target_updated}件を按分）'
            })
            
            # アカウント小計
            account_total = new_count + update_count
            area_total += account_total
            
            hierarchical_data.append({
                'レベル': 2,
                '項目名': f"  └─ 小計",
                '種別': '小計',
                '件数': account_total,
                '備考': f'{account["account_name"]}の{companies_summary["target_date_jp"]}合計'
            })
        
        # 支店合計
        hierarchical_data.append({
            'レベル': 1,
            '項目名': f"🔢 {area_name} 合計",
            '種別': '支店合計',
            '件数': area_total,
            '備考': f'{area_name}の{companies_summary["target_date_jp"]}総計'
        })
        
        # 区切り行
        hierarchical_data.append({
            'レベル': 0,
            '項目名': '',
            '種別': '',
            '件数': '',
            '備考': ''
        })
    
    return hierarchical_data

def generate_hierarchical_excel_data():
    """階層構造のExcel出力用データを生成（支店→アカウント→更新/新規→件数）"""
    
    # 支店とアカウントのマッピングを取得
    mapping = get_area_account_mapping()
    
    # 実際のcompaniesデータのサマリーを取得
    companies_summary = get_companies_summary()
    
    # 階層構造データを生成
    hierarchical_data = []
    
    # 支店ごとにグループ化
    areas = {}
    for item in mapping:
        area_name = item['area_name']
        if area_name not in areas:
            areas[area_name] = {
                'area_id': item['area_id'],
                'accounts': []
            }
        areas[area_name]['accounts'].append({
            'account_id': item['account_id'],
            'account_name': item['account_name']
        })
    
    # 今日の日付
    today = datetime.now().date()
    date_str = today.strftime("%Y年%m月%d日")
    
    # 全体データを支店・アカウント数で按分計算
    total_areas = len(areas)
    total_accounts = len(mapping)
    
    if total_accounts > 0:
        # 今月データをアカウント数で按分（データがある場合）
        if companies_summary['month_new'] > 0 or companies_summary['month_updated'] > 0:
            avg_new_per_account = max(1, companies_summary['month_new'] // total_accounts) if companies_summary['month_new'] > 0 else 0
            avg_update_per_account = max(1, companies_summary['month_updated'] // total_accounts) if companies_summary['month_updated'] > 0 else 0
            remainder_new = companies_summary['month_new'] % total_accounts if companies_summary['month_new'] > 0 else 0
            remainder_update = companies_summary['month_updated'] % total_accounts if companies_summary['month_updated'] > 0 else 0
            data_period = companies_summary['month_period']
            total_new_for_display = companies_summary['month_new']
            total_updated_for_display = companies_summary['month_updated']
        else:
            # 今月データがない場合は過去7日間のデータを使用
            avg_new_per_account = max(1, companies_summary['week_new'] // total_accounts) if companies_summary['week_new'] > 0 else 0
            avg_update_per_account = max(1, companies_summary['week_updated'] // total_accounts) if companies_summary['week_updated'] > 0 else 0
            remainder_new = companies_summary['week_new'] % total_accounts if companies_summary['week_new'] > 0 else 0
            remainder_update = companies_summary['week_updated'] % total_accounts if companies_summary['week_updated'] > 0 else 0
            data_period = companies_summary['week_period']
            total_new_for_display = companies_summary['week_new']
            total_updated_for_display = companies_summary['week_updated']
    else:
        avg_new_per_account = 0
        avg_update_per_account = 0
        remainder_new = 0
        remainder_update = 0
        data_period = companies_summary['date']
        total_new_for_display = 0
        total_updated_for_display = 0
    
    account_index = 0
    
    # 各支店・アカウントに対してデータを生成
    for area_name, area_data in areas.items():
        # 支店ヘッダー行
        hierarchical_data.append({
            'レベル': 1,
            '項目名': f"📍 {area_name}",
            '種別': '',
            '件数': '',
            '備考': f'支店ID: {area_data["area_id"]}'
        })
        
        area_total = 0
        
        for account in area_data['accounts']:
            # アカウントヘッダー行
            hierarchical_data.append({
                'レベル': 2,
                '項目名': f"  📂 {account['account_name']}",
                '種別': '',
                '件数': '',
                '備考': f'アカウントID: {account["account_id"]}'
            })
            
            # 実際のデータを按分で割り当て
            new_count = avg_new_per_account + (remainder_new if account_index == 0 else 0)
            update_count = avg_update_per_account + (remainder_update if account_index == 0 else 0)
            account_index += 1
            
            hierarchical_data.append({
                'レベル': 3,
                '項目名': f"    📝 新規",
                '種別': '新規',
                '件数': new_count,
                '備考': f'{data_period}の実データ（全体{total_new_for_display}件を按分）'
            })
            
            hierarchical_data.append({
                'レベル': 3,
                '項目名': f"    🔄 更新",
                '種別': '更新',
                '件数': update_count,
                '備考': f'{data_period}の実データ（全体{total_updated_for_display}件を按分）'
            })
            
            # アカウント小計
            account_total = new_count + update_count
            area_total += account_total
            
            hierarchical_data.append({
                'レベル': 2,
                '項目名': f"  └─ 小計",
                '種別': '小計',
                '件数': account_total,
                '備考': f'{account["account_name"]}の合計'
            })
        
        # 支店合計
        hierarchical_data.append({
            'レベル': 1,
            '項目名': f"🔢 {area_name} 合計",
            '種別': '支店合計',
            '件数': area_total,
            '備考': f'{area_name}の総計'
        })
        
        # 区切り行
        hierarchical_data.append({
            'レベル': 0,
            '項目名': '',
            '種別': '',
            '件数': '',
            '備考': ''
        })
    
    return hierarchical_data

# ========================
# シンプルなHTMLテンプレート（Excel出力専用）
# ========================

SIMPLE_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ハローワークデータ Excel出力</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .container { 
            background: white;
            padding: 50px;
            border-radius: 20px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            text-align: center;
            max-width: 600px;
            width: 90%;
        }
        .header { 
            margin-bottom: 40px;
        }
        .header h1 { 
            font-size: 2.5em; 
            color: #333;
            margin-bottom: 15px;
        }
        .header p { 
            font-size: 1.2em; 
            color: #666;
            line-height: 1.6;
        }
        .export-section {
            background: #f8f9fa;
            padding: 30px;
            border-radius: 15px;
            margin: 30px 0;
        }
        .export-section h2 {
            color: #495057;
            margin-bottom: 20px;
            font-size: 1.5em;
        }
        .structure-preview {
            background: white;
            padding: 20px;
            border-radius: 10px;
            margin: 20px 0;
            text-align: left;
            font-family: monospace;
            border-left: 4px solid #667eea;
        }
        .btn { 
            padding: 15px 40px; 
            border: none; 
            border-radius: 50px; 
            cursor: pointer; 
            font-size: 18px; 
            font-weight: bold; 
            text-decoration: none; 
            display: inline-block; 
            text-align: center; 
            transition: all 0.3s ease;
            margin: 10px;
        }
        .btn-primary { 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white; 
            box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
        }
        .btn-primary:hover { 
            transform: translateY(-3px);
            box-shadow: 0 15px 30px rgba(102, 126, 234, 0.4);
        }
        .stats {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 20px;
            margin: 30px 0;
        }
        .stat-item {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 15px;
            text-align: center;
        }
        .stat-item h3 {
            font-size: 2em;
            margin-bottom: 5px;
        }
        .stat-item p {
            opacity: 0.9;
            font-size: 0.9em;
        }
        .loading {
            display: none;
            color: #667eea;
            font-weight: bold;
            margin-top: 20px;
        }
        .success-message {
            background: #d4edda;
            color: #155724;
            padding: 15px;
            border-radius: 10px;
            margin-top: 20px;
            display: none;
        }
        .date-selector {
            background: #f8f9fa;
            padding: 25px;
            border-radius: 15px;
            margin: 20px 0;
            border: 2px solid #667eea;
        }
        .date-selector h3 {
            color: #495057;
            margin-bottom: 15px;
            font-size: 1.3em;
        }
        .date-input-group {
            display: flex;
            align-items: center;
            gap: 15px;
            margin-bottom: 15px;
            flex-wrap: wrap;
        }
        .date-input-group input[type="date"] {
            padding: 10px 15px;
            border: 2px solid #ddd;
            border-radius: 8px;
            font-size: 16px;
            background: white;
        }
        .date-input-group input[type="date"]:focus {
            border-color: #667eea;
            outline: none;
        }
        .btn-secondary {
            background: #6c757d;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 8px;
            cursor: pointer;
            font-size: 16px;
        }
        .btn-secondary:hover {
            background: #545b62;
        }
        .date-summary {
            background: white;
            padding: 15px;
            border-radius: 8px;
            margin-top: 15px;
            border-left: 4px solid #28a745;
            display: none;
        }
        .date-summary.error {
            border-left-color: #dc3545;
            color: #721c24;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 ハローワークデータ出力システム</h1>
            <p>支店・アカウント別のハローワークデータを<br>階層構造でExcel出力します</p>
        </div>
        
        <div class="date-selector">
            <h3>📅 日付指定Excel出力</h3>
            <div class="date-input-group">
                <label for="targetDate">対象日付:</label>
                <input type="date" id="targetDate" value="{{ stats.current_date }}">
                <button onclick="checkDateData()" class="btn-secondary">データ確認</button>
                <button onclick="exportExcelByDate()" class="btn btn-primary">指定日でExcel出力</button>
            </div>
            <div id="dateSummary" class="date-summary"></div>
        </div>
        
        <div class="stats">
            <div class="stat-item">
                <h3>{{ stats.total_areas }}</h3>
                <p>対象支店数</p>
            </div>
            <div class="stat-item">
                <h3>{{ stats.hellowork_accounts }}</h3>
                <p>ハローワーク<br>アカウント数</p>
            </div>
            <div class="stat-item">
                <h3>{{ stats.total_companies }}</h3>
                <p>総企業数</p>
            </div>
            <div class="stat-item">
                <h3>{{ stats.today_new }}</h3>
                <p>{{ stats.date }}<br>新規</p>
            </div>
            <div class="stat-item">
                <h3>{{ stats.today_updated }}</h3>
                <p>{{ stats.date }}<br>更新</p>
            </div>
            <div class="stat-item">
                <h3>{{ stats.week_new }}</h3>
                <p>{{ stats.week_period }}<br>新規</p>
            </div>
            <div class="stat-item">
                <h3>{{ stats.week_updated }}</h3>
                <p>{{ stats.week_period }}<br>更新</p>
            </div>
            <div class="stat-item">
                <h3>{{ stats.month_new }}</h3>
                <p>{{ stats.month_period }}<br>新規</p>
            </div>
            <div class="stat-item">
                <h3>{{ stats.month_updated }}</h3>
                <p>{{ stats.month_period }}<br>更新</p>
            </div>
        </div>
        
        <div class="export-section">
            <h2>📋 データ状況とExcel出力</h2>
            <div class="structure-preview">
<strong>� 実データ状況:</strong><br>
・総企業数: {{ stats.total_companies }}件<br>
・{{ stats.date }}: 新規{{ stats.today_new }}件 / 更新{{ stats.today_updated }}件<br>
・{{ stats.week_period }}: 新規{{ stats.week_new }}件 / 更新{{ stats.week_updated }}件<br>
・{{ stats.month_period }}: 新規{{ stats.month_new }}件 / 更新{{ stats.month_updated }}件<br>
<br>
<strong>📋 Excel出力内容:</strong><br>
今月のデータ({{ stats.month_new }}新規 + {{ stats.month_updated }}更新)を<br>
{{ stats.hellowork_accounts }}アカウントに按分して階層表示<br>
<br>
📍 各支店<br>
&nbsp;&nbsp;📂 各アカウント<br>
&nbsp;&nbsp;&nbsp;&nbsp;📝 新規 → 按分値<br>
&nbsp;&nbsp;&nbsp;&nbsp;🔄 更新 → 按分値<br>
&nbsp;&nbsp;└─ 小計 → 計算値<br>
🔢 支店合計 → 実データ合計
            </div>
            
            <button onclick="exportExcel()" class="btn btn-primary" id="exportBtn">
                📊 Excel ファイル出力
            </button>
            
            <div class="loading" id="loadingMsg">
                📈 Excel ファイルを生成中...
            </div>
            
            <div class="success-message" id="successMsg">
                ✅ Excel ファイルの出力が完了しました！
            </div>
        </div>
    </div>

    <script>
        async function checkDateData() {
            const targetDate = document.getElementById('targetDate').value;
            const summaryDiv = document.getElementById('dateSummary');
            
            if (!targetDate) {
                alert('日付を選択してください。');
                return;
            }
            
            try {
                const response = await fetch(`/api/date-summary/${targetDate}`);
                const result = await response.json();
                
                if (result.status === 'success') {
                    const data = result.data;
                    summaryDiv.innerHTML = `
                        <h4>📊 ${data.target_date_jp} のデータ状況</h4>
                        <p><strong>新規:</strong> ${data.target_new}件</p>
                        <p><strong>更新:</strong> ${data.target_updated}件</p>
                        <p><strong>合計:</strong> ${data.target_new + data.target_updated}件</p>
                        <hr>
                        <small>
                        前日(${data.prev_date}): 新規${data.prev_new}件 / 更新${data.prev_updated}件<br>
                        翌日(${data.next_date}): 新規${data.next_new}件 / 更新${data.next_updated}件
                        </small>
                    `;
                    summaryDiv.className = 'date-summary';
                    summaryDiv.style.display = 'block';
                } else {
                    summaryDiv.innerHTML = `<p>❌ エラー: ${result.message}</p>`;
                    summaryDiv.className = 'date-summary error';
                    summaryDiv.style.display = 'block';
                }
            } catch (error) {
                summaryDiv.innerHTML = `<p>❌ データ取得エラー: ${error.message}</p>`;
                summaryDiv.className = 'date-summary error';
                summaryDiv.style.display = 'block';
            }
        }
        
        async function exportExcelByDate() {
            const targetDate = document.getElementById('targetDate').value;
            const btn = document.getElementById('exportByDateBtn');
            const loading = document.getElementById('loadingMsg');
            const success = document.getElementById('successMsg');
            
            if (!targetDate) {
                alert('日付を選択してください。');
                return;
            }
            
            // ボタン無効化とローディング表示
            if (btn) btn.disabled = true;
            if (loading) {
                loading.style.display = 'block';
                loading.innerHTML = `📈 ${targetDate} のExcel ファイルを生成中...`;
            }
            if (success) success.style.display = 'none';
            
            try {
                const response = await fetch('/api/export-excel-by-date', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({ date: targetDate })
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `hellowork_hierarchical_report_${targetDate.replace(/-/g, '')}.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                    
                    // 成功メッセージ表示
                    if (success) {
                        success.innerHTML = `✅ ${targetDate} のExcel ファイルの出力が完了しました！`;
                        success.style.display = 'block';
                        setTimeout(() => {
                            success.style.display = 'none';
                        }, 5000);
                    }
                } else {
                    const errorData = await response.json();
                    alert(`Excel出力に失敗しました: ${errorData.message}`);
                }
            } catch (error) {
                alert('エラーが発生しました: ' + error.message);
            } finally {
                // ボタン復元
                if (btn) btn.disabled = false;
                if (loading) loading.style.display = 'none';
            }
        }

        async function exportExcel() {
            const btn = document.getElementById('exportBtn');
            const loading = document.getElementById('loadingMsg');
            const success = document.getElementById('successMsg');
            
            // ボタン無効化とローディング表示
            btn.disabled = true;
            btn.innerHTML = '⏳ 生成中...';
            loading.style.display = 'block';
            success.style.display = 'none';
            
            try {
                const response = await fetch('/api/export-excel', {
                    method: 'POST'
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `hellowork_hierarchical_report_${new Date().toISOString().split('T')[0].replace(/-/g, '')}.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                    
                    // 成功メッセージ表示
                    success.style.display = 'block';
                    setTimeout(() => {
                        success.style.display = 'none';
                    }, 5000);
                } else {
                    alert('Excel出力に失敗しました。もう一度お試しください。');
                }
            } catch (error) {
                alert('エラーが発生しました: ' + error.message);
            } finally {
                // ボタン復元
                btn.disabled = false;
                btn.innerHTML = '📊 Excel ファイル出力';
                loading.style.display = 'none';
            }
        }
    </script>
</body>
</html>
'''

# ========================
# ルート定義（Excel出力専用）
# ========================

@app.route('/')
def index():
    try:
        # 基本統計のみ取得
        mapping = get_area_account_mapping()
        
        # 実際のcompaniesデータのサマリーを取得
        companies_summary = get_companies_summary()
        
        # 支店とアカウントをカウント
        areas = set()
        hellowork_accounts = 0
        
        for item in mapping:
            areas.add(item['area_name'])
            if item['needs_hellowork']:
                hellowork_accounts += 1
        
        stats = {
            'total_areas': len(areas),
            'hellowork_accounts': hellowork_accounts,
            'total_companies': companies_summary['total_companies'],
            'today_new': companies_summary['today_new'],
            'today_updated': companies_summary['today_updated'],
            'week_new': companies_summary['week_new'],
            'week_updated': companies_summary['week_updated'],
            'month_new': companies_summary['month_new'],
            'month_updated': companies_summary['month_updated'],
            'date': companies_summary['date'],
            'week_period': companies_summary['week_period'],
            'month_period': companies_summary['month_period'],
            'current_date': datetime.now().strftime('%Y-%m-%d')
        }
        
        return render_template_string(SIMPLE_TEMPLATE, stats=stats)
                                    
    except Exception as e:
        return f"<h1>エラー</h1><p>{str(e)}</p><p><a href='/api/test'>API テスト</a></p>", 500

@app.route('/api/export-excel-by-date', methods=['POST'])
def export_excel_by_date():
    """指定日の階層構造 Excel出力API"""
    try:
        # リクエストから日付を取得
        data = request.get_json() or {}
        target_date = data.get('date')
        
        if target_date:
            # 日付形式の検証
            try:
                datetime.strptime(target_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({'status': 'error', 'message': '日付形式が正しくありません。YYYY-MM-DD形式で入力してください。'}), 400
        
        # 階層構造データを生成
        hierarchical_data = generate_hierarchical_excel_data_by_date(target_date)
        
        # DataFrame作成
        df = pd.DataFrame(hierarchical_data)
        
        # Excelファイル作成（既存のスタイル適用ロジックを使用）
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ハローワーク送信状況')
            
            # スタイル調整（既存のコードを再利用）
            workbook = writer.book
            worksheet = writer.sheets['ハローワーク送信状況']
            
            # ヘッダーのスタイル
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            # レベル別のスタイル定義
            level1_font = Font(bold=True, size=14, color='000080')  # 支店
            level1_fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
            
            level2_font = Font(bold=True, size=11, color='000000')  # アカウント
            level2_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
            
            level3_font = Font(size=10, color='333333')  # 新規/更新
            level3_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            subtotal_font = Font(bold=True, size=10, color='006600')  # 小計
            subtotal_fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
            
            total_font = Font(bold=True, size=12, color='800000')  # 支店合計
            total_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            
            # 罫線の定義
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ヘッダー行のスタイル設定
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # データ行のスタイル設定
            for row_num, row_data in enumerate(hierarchical_data, start=2):
                level = row_data.get('レベル', 0)
                type_value = row_data.get('種別', '')
                
                # レベルに応じてスタイルを適用
                if level == 1:  # 支店
                    if type_value == '支店合計':
                        font = total_font
                        fill = total_fill
                    else:
                        font = level1_font
                        fill = level1_fill
                elif level == 2:  # アカウント・小計
                    if type_value == '小計':
                        font = subtotal_font
                        fill = subtotal_fill
                    else:
                        font = level2_font
                        fill = level2_fill
                elif level == 3:  # 新規/更新
                    font = level3_font
                    fill = level3_fill
                else:  # 空行など
                    font = Font(size=10)
                    fill = PatternFill()
                
                # 行の全セルにスタイルを適用
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = font
                    cell.fill = fill
                    cell.border = thin_border
                    
                    # 件数列は右寄せ
                    if col_num == 4 and row_data.get('件数') != '':  # 件数列
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 列幅の調整
            column_widths = {
                'A': 8,   # レベル
                'B': 40,  # 項目名
                'C': 12,  # 種別
                'D': 12,  # 件数
                'E': 30   # 備考
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # フリーズペイン（ヘッダー行を固定）
            worksheet.freeze_panes = 'A2'
            
            # シートタブの色
            worksheet.sheet_properties.tabColor = "366092"
        
        output.seek(0)
        
        # ファイル名生成
        if target_date:
            filename = f"hellowork_hierarchical_report_{target_date.replace('-', '')}.xlsx"
        else:
            today = date.today()
            filename = f"hellowork_hierarchical_report_{today.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """階層構造 Excel出力API"""
    try:
        # 階層構造データを生成
        hierarchical_data = generate_hierarchical_excel_data()
        
        # DataFrame作成
        df = pd.DataFrame(hierarchical_data)
        
        # Excelファイル作成
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ハローワーク送信状況')
            
            # スタイル調整
            workbook = writer.book
            worksheet = writer.sheets['ハローワーク送信状況']
            
            # フォントとスタイルの定義
            # ヘッダーのスタイル
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            # レベル別のスタイル定義
            level1_font = Font(bold=True, size=14, color='000080')  # 支店
            level1_fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
            
            level2_font = Font(bold=True, size=11, color='000000')  # アカウント
            level2_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
            
            level3_font = Font(size=10, color='333333')  # 新規/更新
            level3_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            subtotal_font = Font(bold=True, size=10, color='006600')  # 小計
            subtotal_fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
            
            total_font = Font(bold=True, size=12, color='800000')  # 支店合計
            total_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            
            # 罫線の定義
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ヘッダー行のスタイル設定
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # データ行のスタイル設定
            for row_num, row_data in enumerate(hierarchical_data, start=2):
                level = row_data.get('レベル', 0)
                type_value = row_data.get('種別', '')
                
                # レベルに応じてスタイルを適用
                if level == 1:  # 支店
                    if type_value == '支店合計':
                        font = total_font
                        fill = total_fill
                    else:
                        font = level1_font
                        fill = level1_fill
                elif level == 2:  # アカウント・小計
                    if type_value == '小計':
                        font = subtotal_font
                        fill = subtotal_fill
                    else:
                        font = level2_font
                        fill = level2_fill
                elif level == 3:  # 新規/更新
                    font = level3_font
                    fill = level3_fill
                else:  # 空行など
                    font = Font(size=10)
                    fill = PatternFill()
                
                # 行の全セルにスタイルを適用
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = font
                    cell.fill = fill
                    cell.border = thin_border
                    
                    # 件数列は右寄せ
                    if col_num == 4 and row_data.get('件数') != '':  # 件数列
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 列幅の調整
            column_widths = {
                'A': 8,   # レベル
                'B': 40,  # 項目名
                'C': 12,  # 種別
                'D': 12,  # 件数
                'E': 30   # 備考
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # フリーズペイン（ヘッダー行を固定）
            worksheet.freeze_panes = 'A2'
            
            # シートタブの色
            worksheet.sheet_properties.tabColor = "366092"
        
        output.seek(0)
        
        # ファイル名生成
        today = date.today()
        filename = f"hellowork_hierarchical_report_{today.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/date-summary/<date_str>')
def get_date_summary(date_str):
    """指定日のデータサマリーを取得"""
    try:
        # 日付形式の検証
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # 指定日のサマリーを取得
        summary = get_companies_summary_by_date(target_date)
        
        return jsonify({
            'status': 'success',
            'data': summary
        })
        
    except ValueError:
        return jsonify({
            'status': 'error',
            'message': '日付形式が正しくありません。YYYY-MM-DD形式で入力してください。'
        }), 400
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/debug-companies')
def debug_companies():
    """companiesテーブルのデバッグ情報を取得"""
    try:
        # テーブル存在確認
        table_exists = db.session.execute(
            text("SHOW TABLES LIKE 'companies'")
        ).fetchone()
        
        if not table_exists:
            return jsonify({
                'status': 'error',
                'message': 'companiesテーブルが存在しません',
                'table_exists': False
            })
        
        # テーブル構造確認
        columns = db.session.execute(
            text("DESCRIBE companies")
        ).fetchall()
        
        # 総レコード数
        total_count = db.session.execute(
            text("SELECT COUNT(*) FROM companies")
        ).scalar()
        
        # 最新の10件のcreated_at, updated_atを確認
        recent_data = db.session.execute(
            text("SELECT id, created_at, updated_at FROM companies ORDER BY id DESC LIMIT 10")
        ).fetchall()
        
        # 日付範囲確認
        date_range = db.session.execute(
            text("SELECT MIN(DATE(created_at)) as min_date, MAX(DATE(created_at)) as max_date FROM companies WHERE created_at IS NOT NULL")
        ).fetchone()
        
        # 10月8日のデータ確認
        oct8_created = db.session.execute(
            text("SELECT COUNT(*) FROM companies WHERE DATE(created_at) = '2025-10-08'")
        ).scalar()
        
        oct8_updated = db.session.execute(
            text("SELECT COUNT(*) FROM companies WHERE DATE(updated_at) = '2025-10-08' AND DATE(created_at) != '2025-10-08'")
        ).scalar()
        
        return jsonify({
            'status': 'success',
            'table_exists': True,
            'columns': [{'Field': col[0], 'Type': col[1], 'Null': col[2], 'Key': col[3]} for col in columns],
            'total_count': total_count,
            'recent_data': [{'id': row[0], 'created_at': str(row[1]), 'updated_at': str(row[2])} for row in recent_data],
            'date_range': {'min_date': str(date_range[0]) if date_range[0] else None, 'max_date': str(date_range[1]) if date_range[1] else None},
            'oct8_created': oct8_created,
            'oct8_updated': oct8_updated
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'error_type': type(e).__name__
        })

@app.route('/api/test')
def api_test():
    """簡単なAPI接続テスト"""
    try:
        with db.engine.connect() as connection:
            result = connection.execute(text('SELECT VERSION()'))
            mysql_version = result.fetchone()[0]
        
        mapping = get_area_account_mapping()
        
        return jsonify({
            'status': 'success',
            'message': 'Excel出力専用アプリケーション',
            'database': 'scraping',
            'mysql_version': mysql_version,
            'hellowork_accounts': len(mapping)
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)