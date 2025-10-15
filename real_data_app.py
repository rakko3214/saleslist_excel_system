from flask import Flask, render_template_string, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, func, and_
from dotenv import load_dotenv
import os
import pymysql
from datetime import datetime, date, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io

# 環境変数をロード
load_dotenv()

app = Flask(__name__)

# データベース設定
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')

db = SQLAlchemy(app)

# ========================
# 実際のデータ構造に合わせたモデル定義
# ========================

class FmArea(db.Model):
    __tablename__ = 'fm_areas'
    
    id = db.Column(db.Integer, primary_key=True)
    area_name_ja = db.Column(db.Text, nullable=False)
    area_name_en = db.Column(db.Text, nullable=False)
    fm_login_account_id = db.Column(db.Text, nullable=False)
    fm_login_account_pass = db.Column(db.Text, nullable=False)
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.area_name_ja,
            'name_en': self.area_name_en,
            'login_id': self.fm_login_account_id
        }

class FmAccount(db.Model):
    __tablename__ = 'fm_accounts'
    
    id = db.Column(db.Integer, primary_key=True)
    department_name = db.Column(db.Text, nullable=False)
    sort_order = db.Column(db.Integer, nullable=False)
    needs_hellowork = db.Column(db.Integer, nullable=False, default=0)
    needs_tabelog = db.Column(db.Integer, nullable=False, default=0)
    needs_kanri = db.Column(db.Integer, nullable=False, default=1)
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.department_name,
            'sort_order': self.sort_order,
            'needs_hellowork': bool(self.needs_hellowork),
            'needs_tabelog': bool(self.needs_tabelog),
            'needs_kanri': bool(self.needs_kanri)
        }

class Company(db.Model):
    __tablename__ = 'companies'
    
    id = db.Column(db.Integer, primary_key=True)
    media_site_id = db.Column(db.Integer)
    fm_area_id = db.Column(db.Integer)
    company_id_in_site = db.Column(db.Text)
    url = db.Column(db.Text)
    company_name = db.Column(db.Text)
    tel = db.Column(db.Text)
    address = db.Column(db.Text)
    hp = db.Column(db.Text)
    ceo_name = db.Column(db.Text)
    capital_stock = db.Column(db.Integer)
    job_detail = db.Column(db.Text)
    fm_major_industry_id = db.Column(db.Integer)
    fm_minor_industry_id = db.Column(db.Integer)
    kanri_regist_history_id = db.Column(db.Integer)
    share_departments = db.Column(db.Text)
    imported_fm_account_id = db.Column(db.Integer)
    display_started_at = db.Column(db.Date)
    created_at = db.Column(db.DateTime)
    updated_at = db.Column(db.DateTime)
    fm_saved_at = db.Column(db.DateTime)
    fm_import_result = db.Column(db.Integer)
    
    def to_dict(self):
        return {
            'id': self.id,
            'company_name': self.company_name,
            'fm_area_id': self.fm_area_id,
            'address': self.address,
            'tel': self.tel,
            'url': self.url,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

# fm_area_accountsの関連テーブル（多対多関係）
class FmAreaAccount(db.Model):
    __tablename__ = 'fm_area_accounts'
    
    id = db.Column(db.Integer, primary_key=True)
    fm_area_id = db.Column(db.Integer, nullable=False)
    fm_account_id = db.Column(db.Integer, nullable=False)
    is_related = db.Column(db.Integer, nullable=False, default=0)
    
    def to_dict(self):
        return {
            'id': self.id,
            'fm_area_id': self.fm_area_id,
            'fm_account_id': self.fm_account_id,
            'is_related': self.is_related
        }

# ========================
# 集計関数（実データ構造対応）
# ========================

def get_area_account_summary():
    """支店・アカウント・データ件数のサマリーを取得"""
    
    # 支店別の企業データ件数を取得
    area_summary = db.session.query(
        FmArea.id,
        FmArea.area_name_ja,
        func.count(Company.id).label('company_count')
    ).outerjoin(
        Company, FmArea.id == Company.fm_area_id
    ).group_by(
        FmArea.id, FmArea.area_name_ja
    ).order_by(FmArea.id).all()
    
    # アカウント別の関連情報
    account_summary = db.session.query(
        FmAccount.id,
        FmAccount.department_name,
        FmAccount.needs_hellowork,
        func.count(FmAreaAccount.id).label('area_count')
    ).outerjoin(
        FmAreaAccount, FmAccount.id == FmAreaAccount.fm_account_id
    ).group_by(
        FmAccount.id
    ).order_by(FmAccount.sort_order).all()
    
    return {
        'areas': [
            {
                'id': row.id,
                'name': row.area_name_ja,
                'company_count': row.company_count
            } for row in area_summary
        ],
        'accounts': [
            {
                'id': row.id,
                'name': row.department_name,
                'needs_hellowork': row.needs_hellowork,
                'area_count': row.area_count
            } for row in account_summary
        ]
    }

def get_area_account_mapping():
    """実際のデータベース構造に基づく支店とアカウントの関連マッピングを取得（ハローワーク制限なし）"""
    
    mapping = db.session.query(
        FmAreaAccount.fm_area_id,
        FmAreaAccount.fm_account_id,
        FmArea.area_name_ja,
        FmAccount.department_name,
        FmAccount.needs_hellowork,
        FmAreaAccount.is_related
    ).join(
        FmArea, FmAreaAccount.fm_area_id == FmArea.id
    ).join(
        FmAccount, FmAreaAccount.fm_account_id == FmAccount.id
    ).filter(
        FmAreaAccount.is_related == 1    # メイン関係のアカウントのみ（ハローワーク制限なし）
    ).order_by(
        FmArea.id, FmAccount.sort_order
    ).all()
    
    return [
        {
            'area_id': row.fm_area_id,
            'area_name': row.area_name_ja,
            'account_id': row.fm_account_id,
            'account_name': row.department_name,
            'needs_hellowork': row.needs_hellowork,
            'is_related': row.is_related
        } for row in mapping
    ]

def get_all_areas_with_accounts():
    """全支店と関連アカウント情報を取得（ハローワーク制限なし、データ存在チェック付き）"""
    
    # 全支店を取得
    all_areas = db.session.query(FmArea.id, FmArea.area_name_ja).order_by(FmArea.id).all()
    
    # アカウントマッピングを取得（ハローワーク制限なし）
    mapping = get_area_account_mapping()
    
    # 支店ごとにグループ化し、実際のデータ存在チェック
    areas_with_accounts = []
    
    for area in all_areas:
        area_id, area_name = area
        
        # この支店に関連するアカウントを取得
        area_accounts = [item for item in mapping if item['area_id'] == area_id]
        
        # この支店に実際のcompaniesデータがあるかチェック
        has_data = False
        if area_accounts:
            try:
                # 支店にデータがあるかチェック
                data_count = db.session.query(Company).filter(
                    Company.fm_area_id == area_id
                ).count()
                has_data = data_count > 0
            except:
                has_data = False
        
        areas_with_accounts.append({
            'area_id': area_id,
            'area_name': area_name,
            'accounts': area_accounts,
            'has_hellowork_accounts': len(area_accounts) > 0,
            'has_data': has_data
        })
    
    return areas_with_accounts

def get_companies_data_by_period(area_id, account_id, date_filter='today', start_date=None, end_date=None):
    """期間指定で企業データを取得（支店・アカウント別）- 軽量化対応"""
    from datetime import datetime, timedelta
    
    # 期間の計算
    today = datetime.now().date()
    
    if date_filter == 'today':
        filter_start = today
        filter_end = today
    elif date_filter == 'week':
        filter_start = today - timedelta(days=7)
        filter_end = today
    elif date_filter == 'month':
        filter_start = today - timedelta(days=30)
        filter_end = today
    elif date_filter == 'year':
        filter_start = today - timedelta(days=365)
        filter_end = today
    elif date_filter == 'custom' and start_date and end_date:
        filter_start = start_date
        filter_end = end_date
    else:
        filter_start = today
        filter_end = today
    
    # 軽量化: 個別クエリではなく一度にまとめて取得
    try:
        # 基本フィルタ
        base_query = db.session.query(Company).filter(
            Company.fm_area_id == area_id,
            Company.imported_fm_account_id == account_id
        )
        
        # 新規データ（fm_import_result = 2）
        new_count = base_query.filter(
            Company.fm_import_result == 2,
            func.date(Company.created_at).between(filter_start, filter_end)
        ).count()
        
        # 更新データ（fm_import_result = 1）
        update_count = base_query.filter(
            Company.fm_import_result == 1,
            func.date(Company.updated_at).between(filter_start, filter_end)
        ).count()
        
        # 軽量化: 支部レベルでは振り分けなしは常に0（計算省略）
        unassigned_count = 0
        
        return {
            'new_count': new_count,
            'update_count': update_count,
            'unassigned_count': unassigned_count,
            'period': f'{filter_start} 〜 {filter_end}'
        }
        
    except Exception as e:
        print(f"データ取得エラー（軽量化）: {e}")
        # エラー時は空データを返す（フォールバック処理を軽量化）
        return {
            'new_count': 0,
            'update_count': 0,
            'unassigned_count': 0,
            'period': f'{filter_start} 〜 {filter_end} (エラー)'
        }

def generate_hierarchical_excel_data(date_filter='today', start_date=None, end_date=None):
    """画像フォーマットに対応した階層構造のExcel出力用データを生成"""
    from datetime import datetime, timedelta
    
    # 全支店と関連アカウント情報を取得
    areas_with_accounts = get_all_areas_with_accounts()
    
    # 階層構造データを生成
    hierarchical_data = []
    
    # 期間情報の表示用
    period_info = ""
    date_text = ""
    if date_filter == 'today':
        period_info = "本日"
        date_text = datetime.now().strftime("%Y年%m月%d日")
    elif date_filter == 'week':
        period_info = "1週間"
        date_text = datetime.now().strftime("%Y年%m月%d日")
    elif date_filter == 'month':
        period_info = "1ヶ月"
        date_text = datetime.now().strftime("%Y年%m月%d日")
    elif date_filter == 'year':
        period_info = "1年"
        date_text = datetime.now().strftime("%Y年%m月%d日")
    elif date_filter == 'custom':
        period_info = f"{start_date}〜{end_date}"
        date_text = f"{start_date}〜{end_date}"
    
    # 全支店のデータを構築（画像フォーマット準拠）
    for area_info in areas_with_accounts:
        # データがある支店のみを処理（ハローワーク制限なし）
        if not area_info['accounts']:
            continue  # アカウントがない支店はスキップ
            
        # 支店ヘッダー行（項目名のみ）
        hierarchical_data.append({
            'レベル': 1,
            '項目名': f'📍 {area_info["area_name"]}',
            '種別': '',
            '件数': '',
            '備考': f'支店ID: {area_info["area_id"]}'
        })
        
        area_total_new = 0
        area_total_update = 0
        area_total_unassigned = 0
        
        for account_info in area_info['accounts']:
            # アカウントヘッダー行（項目名のみ）
            hierarchical_data.append({
                'レベル': 2,
                '項目名': f'├─ {account_info["account_name"]}',
                '種別': '',
                '件数': '',
                '備考': f'アカウントID: {account_info["account_id"]}'
            })
            
            # 実際のデータベースから件数を取得
            try:
                data_result = get_companies_data_by_period(
                    area_info["area_id"], 
                    account_info["account_id"],
                    date_filter=date_filter,
                    start_date=start_date,
                    end_date=end_date
                )
                
                new_count = data_result['new_count']
                update_count = data_result['update_count']
                unassigned_count = data_result['unassigned_count']
                # 部門レベルは振り分けなしを除外した合計
                total_count = new_count + update_count
                
                # 新規データ行（画像フォーマット）
                hierarchical_data.append({
                    'レベル': 3,
                    '項目名': '│  ├─ 新規',
                    '種別': '新規',
                    '件数': new_count,
                    '備考': f'{date_text}の全データ（全体SO件対象分）'
                })
                
                # 更新データ行（画像フォーマット）
                hierarchical_data.append({
                    'レベル': 3,
                    '項目名': '│  ├─ 更新',
                    '種別': '更新',
                    '件数': update_count,
                    '備考': f'{date_text}の全データ（全体2365件対象分）'
                })
                
                # 振り分けなしデータ行（画像フォーマット）
                hierarchical_data.append({
                    'レベル': 3,
                    '項目名': '│  ├─ 振り分けなし',
                    '種別': '振り分けなし',
                    '件数': unassigned_count,
                    '備考': f'{date_text}の全データ（振り分けなし分）'
                })
                
                # アカウント小計行（画像フォーマット）
                hierarchical_data.append({
                    'レベル': 3,
                    '項目名': '│  └─ 小計',
                    '種別': '小計',
                    '件数': total_count,
                    '備考': f'{account_info["account_name"]}の{date_text}合計'
                })
                
                area_total_new += new_count
                area_total_update += update_count
                area_total_unassigned += unassigned_count
                
            except Exception as e:
                print(f"データ取得エラー: 支店{area_info['area_id']}, アカウント{account_info['account_id']}: {e}")
                # エラー時は0で補完
                hierarchical_data.append({
                    'レベル': 3,
                    '項目名': '│  ├─ 新規',
                    '種別': '新規',
                    '件数': 0,
                    '備考': f'{date_text}の全データ（エラー）'
                })
                
                hierarchical_data.append({
                    'レベル': 3,
                    '項目名': '│  ├─ 更新',
                    '種別': '更新',
                    '件数': 0,
                    '備考': f'{date_text}の全データ（エラー）'
                })
                
                hierarchical_data.append({
                    'レベル': 3,
                    '項目名': '│  ├─ 振り分けなし',
                    '種別': '振り分けなし',
                    '件数': 0,
                    '備考': f'{date_text}の全データ（エラー）'
                })
                
                hierarchical_data.append({
                    'レベル': 3,
                    '項目名': '│  └─ 小計',
                    '種別': '小計',
                    '件数': 0,
                    '備考': f'{account_info["account_name"]}の{date_text}合計（エラー）'
                })
        
        # 支店合計行（画像フォーマット）
        area_total = area_total_new + area_total_update + area_total_unassigned
        hierarchical_data.append({
            'レベル': 1,
            '項目名': f'└─ {area_info["area_name"]} 合計',
            '種別': '支店合計',
            '件数': area_total,
            '備考': f'{area_info["area_name"]}の{date_text}総計'
        })
        
        # 支店間の区切り行
        hierarchical_data.append({
            'レベル': 0,
            '項目名': '',
            '種別': '',
            '件数': '',
            '備考': ''
        })
    
    return hierarchical_data

# ========================
# HTMLテンプレート（実データ対応）
# ========================

MAIN_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ハローワーク営業リスト</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; padding: 20px; }
        .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; }
        .header h1 { font-size: 2.5em; margin-bottom: 10px; }
        .header p { font-size: 1.1em; opacity: 0.9; }
        .card { background: white; border-radius: 10px; padding: 25px; margin-bottom: 20px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
        .card h2 { color: #333; margin-bottom: 20px; padding-bottom: 10px; border-bottom: 3px solid #667eea; }
        .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px; }
        .stat-card { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 10px; text-align: center; }
        .stat-card h3 { font-size: 2em; margin-bottom: 5px; }
        .stat-card p { opacity: 0.9; }
        .table-container { overflow-x: auto; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background-color: #f8f9fa; font-weight: bold; color: #495057; }
        tr:hover { background-color: #f8f9fa; }
        .btn { padding: 12px 25px; border: none; border-radius: 5px; cursor: pointer; font-size: 14px; font-weight: bold; text-decoration: none; display: inline-block; text-align: center; transition: all 0.3s; margin: 5px; }
        .btn-primary { background: #667eea; color: white; }
        .btn-primary:hover { background: #5a6fd8; transform: translateY(-2px); }
        .btn-success { background: #28a745; color: white; }
        .btn-success:hover { background: #218838; transform: translateY(-2px); }
        .btn-info { background: #17a2b8; color: white; }
        .btn-info:hover { background: #138496; transform: translateY(-2px); }
        .status-success { color: #28a745; font-weight: bold; }
        .status-error { color: #dc3545; font-weight: bold; }
        .hellowork-enabled { background-color: #d4edda; }
        .area-section { margin-bottom: 30px; border-left: 4px solid #667eea; padding-left: 20px; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🏢 ハローワーク営業リスト</h1>
            <p>{{ stats.total_companies }}件の企業データ・{{ stats.total_areas }}支店・{{ stats.total_accounts }}アカウントを管理</p>
        </div>

        
        <div class="card">
            <h2>📅 データ表示期間設定</h2>
            <div style="margin-bottom: 20px;">
                <p style="color: #666; margin-bottom: 15px;">
                    📊 <strong>今日:</strong> 5,101件 | <strong>1週間:</strong> 35,864件 | <strong>1ヶ月:</strong> 102,077件 | <strong>全体:</strong> 約72万件
                </p>
                
                <label for="dataFilter" style="font-weight: bold; margin-right: 10px;">表示期間:</label>
                <select id="dataFilter" style="padding: 8px; margin-right: 15px; border: 1px solid #ddd; border-radius: 4px;">
                    <option value="today" selected>今日のデータ (約5千件)</option>
                    <option value="week">1週間 (約3.6万件)</option>
                    <option value="month">1ヶ月 (約10万件)</option>
                    <option value="all">全データ (約72万件) ⚠️重い</option>
                </select>
                
                <button onclick="loadDataWithFilter()" class="btn btn-primary">
                    📊 データを読み込み
                </button>
            </div>
            
            <div id="loadingIndicator" style="display: none; text-align: center; padding: 20px; color: #666;">
                ⏳ データを読み込んでいます...
            </div>
            
            <div id="dataStats" style="padding: 15px; background-color: #e7f3ff; border-radius: 5px; margin-bottom: 15px;">
                <strong>📈 現在表示中:</strong> <span id="currentPeriod">今日のデータ</span> | 
                <strong>件数:</strong> <span id="currentCount">読み込み中...</span>
            </div>
        </div>   
        <div class="card">
            <h2>📅 日付指定データ操作</h2>
            <div style="margin-bottom: 20px;">
                <p style="color: #666; margin-bottom: 15px;">
                    🎯 特定の日付範囲でデータを表示・Excel出力できます
                </p>
                
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px;">
                    <div>
                        <label for="startDate" style="font-weight: bold; display: block; margin-bottom: 5px;">開始日:</label>
                        <input type="date" id="startDate" style="padding: 8px; width: 100%; border: 1px solid #ddd; border-radius: 4px;">
                    </div>
                    <div>
                        <label for="endDate" style="font-weight: bold; display: block; margin-bottom: 5px;">終了日:</label>
                        <input type="date" id="endDate" style="padding: 8px; width: 100%; border: 1px solid #ddd; border-radius: 4px;">
                    </div>
                </div>
                
                <div style="display: flex; gap: 10px; flex-wrap: wrap;">
                    <button onclick="setDateRange('today')" class="btn btn-info">
                        📅 今日
                    </button>
                    <button onclick="setDateRange('week')" class="btn btn-info">
                        📅 1週間
                    </button>
                    <button onclick="setDateRange('month')" class="btn btn-info">
                        📅 1ヶ月
                    </button>
                    <button onclick="loadDataByDateRange()" class="btn btn-primary">
                        📊 日付範囲でデータ表示
                    </button>
                    <button onclick="exportExcelByDateRange()" class="btn btn-success">
                        📋 日付範囲でExcel出力
                    </button>
                </div>
            </div>
            
            <!-- 日付指定データ表示エリア -->
            <div id="dateRangeResults" style="display: none; margin-top: 20px; padding: 20px; background-color: #f8f9fa; border-radius: 5px;">
                <h3 id="dateRangeTitle">日付範囲データ</h3>
                <div id="dateRangeContent">
                    <!-- ここに日付範囲指定データが表示されます -->
                </div>
            </div>
        </div>
        
        <div class="card">
            <h2 style="cursor: pointer; user-select: none;" onclick="toggleAccordion('branchDataSection')">
                🏢 支店別データ 
                <span id="branchToggleIcon" style="float: right; font-size: 1.2em;">▼</span>
            </h2>
            <div id="branchDataSection" style="display: none; margin-top: 15px;">
                <div class="table-container">
                    <table id="areaTable">
                        <thead>
                            <tr>
                                <th>支店ID</th>
                                <th>支店名</th>
                                <th>企業データ件数</th>
                                <th>状況</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for area in areas %}
                            <tr>
                                <td>{{ area.id }}</td>
                                <td><strong>{{ area.name }}</strong></td>
                                <td>{{ area.company_count }}件</td>
                                <td><span class="status-success">✅ 稼働中</span></td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <script>
        async function exportHierarchicalReport() {
            try {
                const response = await fetch('/api/export-mapping', {
                    method: 'POST'
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `hellowork_hierarchical_report_${new Date().toISOString().split('T')[0].replace(/-/g, '')}.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                    alert('階層構造レポートのExcel出力が完了しました！\\n\\n構造:\\n支店 → アカウント → 新規/更新 → 件数');
                } else {
                    alert('Excel出力に失敗しました');
                }
            } catch (error) {
                alert('Excel出力でエラーが発生しました: ' + error.message);
            }
        }

        // 期間指定Excel出力機能
        async function exportWithFilter() {
            const dateFilter = document.getElementById('dateFilter').value;
            const startDate = document.getElementById('startDate').value;
            const endDate = document.getElementById('endDate').value;
            
            // カスタム期間の場合は日付チェック
            if (dateFilter === 'custom' && (!startDate || !endDate)) {
                alert('カスタム期間を選択した場合は、開始日と終了日を入力してください。');
                return;
            }
            
            const requestData = {
                date_filter: dateFilter,
                start_date: startDate || null,
                end_date: endDate || null
            };
            
            try {
                const response = await fetch('/api/export-mapping', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify(requestData)
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `hellowork_data_${dateFilter}_${new Date().toISOString().split('T')[0]}.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                    
                    let periodText = '';
                    switch(dateFilter) {
                        case 'today': periodText = '今日'; break;
                        case 'week': periodText = '1週間'; break;
                        case 'month': periodText = '1ヶ月'; break;
                        case 'year': periodText = '1年'; break;
                        case 'custom': periodText = `${startDate}〜${endDate}`; break;
                    }
                    alert(`${periodText}のデータでExcel出力が完了しました！\\n\\n実際のデータベースから取得したデータです。`);
                } else {
                    alert('Excel出力に失敗しました');
                }
            } catch (error) {
                alert('Excel出力でエラーが発生しました: ' + error.message);
            }
        }
        
        // 期間選択の表示制御（要素存在チェック付き）
        const dateFilterElement = document.getElementById('dateFilter');
        if (dateFilterElement) {
            dateFilterElement.addEventListener('change', function() {
                const customRange = document.getElementById('customDateRange');
                if (customRange && this.value === 'custom') {
                    customRange.style.display = 'block';
                } else if (customRange) {
                    customRange.style.display = 'none';
                }
            });
        }

        // 期間フィルタ機能
        async function loadDataWithFilter() {
            const dataFilter = document.getElementById('dataFilter').value;
            const loadingIndicator = document.getElementById('loadingIndicator');
            const currentPeriod = document.getElementById('currentPeriod');
            const currentCount = document.getElementById('currentCount');
            
            // ローディング表示
            loadingIndicator.style.display = 'block';
            
            try {
                const response = await fetch('/api/filtered-data', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        date_filter: dataFilter
                    })
                });
                
                if (response.ok) {
                    const data = await response.json();
                    
                    if (data.status === 'success') {
                        // 統計情報を更新
                        currentPeriod.textContent = data.period + 'のデータ';
                        currentCount.innerHTML = `
                            合計: ${data.total_companies.toLocaleString()}件
                            <span style="margin-left: 15px; color: #4CAF50; font-weight: bold;">新規: ${data.total_new}件</span>
                            <span style="margin-left: 10px; color: #2196F3;">更新: ${data.total_update}件</span>
                            <span style="margin-left: 10px; color: #FF9800;">振り分けなし: ${data.total_unassigned}件</span>
                        `;
                        
                        // 支店別データテーブルを更新（階層構造）
                        updateAreaTable(data.areas);
                        
                        // マッピング情報を更新
                        updateMappingSection(data.areas);
                        
                        //alert(`✅ ${data.period_text}のデータを読み込みました（合計: ${data.total_companies.toLocaleString()}件, 新規: ${data.total_new}件, 更新: ${data.total_update}件）`);
                    } else {
                        alert('❌ データ読み込みエラー: ' + data.message);
                    }
                } else {
                    alert('❌ サーバーエラーが発生しました');
                }
            } catch (error) {
                alert('❌ データ読み込みでエラーが発生しました: ' + error.message);
            } finally {
                loadingIndicator.style.display = 'none';
            }
        }
        
        // 現在の設定で再読み込み
        function loadCurrentData() {
            loadDataWithFilter();
        }
        
        // テーブル更新関数（データベース構造に忠実な階層表示）
        function updateAreaTable(areas) {
            const tbody = document.querySelector('#areaTable tbody');
            if (tbody) {
                tbody.innerHTML = '';
                
                areas.forEach(area => {
                    // 支店ヘッダー行
                    const areaRow = tbody.insertRow();
                    areaRow.classList.add('area-header');
                    
                    if (area.accounts && area.accounts.length > 0) {
                        // アカウントがある支店
                        areaRow.innerHTML = `
                            <td colspan="4" style="background-color: #e8f4fd; font-weight: bold; padding: 12px;">
                                📍 ${area.name} (ID: ${area.id})
                                <span style="float: right;">
                                    新規: ${area.new_count}件 | 更新: ${area.update_count}件 | 振り分けなし: ${area.unassigned_count}件 | 合計: ${area.total_count}件
                                </span>
                            </td>
                        `;
                        
                        // アカウント詳細行
                        area.accounts.forEach((account, index) => {
                            const accountRow = tbody.insertRow();
                            accountRow.classList.add('account-detail');
                            
                            const isLast = index === area.accounts.length - 1;
                            const treeChar = isLast ? '└─' : '├─';
                            
                            // needs_helloworkの値に応じて表示を切り替え
                            let helloworkBadge;
                            switch(account.needs_hellowork) {
                                case 0:
                                    helloworkBadge = '<span style="background: #DC3545; color: white; padding: 2px 6px; border-radius: 3px; font-size: 11px;">営業なし</span>';
                                    break;
                                case 1:
                                    helloworkBadge = '<span style="background: #00FF00; color: black; padding: 2px 6px; border-radius: 3px; font-size: 11px;">WEBなし</span>';
                                    break;
                                case 2:
                                    helloworkBadge = '<span style="background: #FFC107; color: white; padding: 2px 6px; border-radius: 3px; font-size: 11px;">WEBあり</span>';
                                    break;
                                case 3:
                                    helloworkBadge = '<span style="background: #007BFF; color: white; padding: 2px 6px; border-radius: 3px; font-size: 11px;">両方対応</span>';
                                    break;
                                default:
                                    helloworkBadge = '<span style="background: #6C757D; color: white; padding: 2px 6px; border-radius: 3px; font-size: 11px;">不明</span>';
                            }
                            
                            accountRow.innerHTML = `
                                <td style="padding-left: 20px;">${treeChar} ${account.name}</td>
                                <td>${helloworkBadge}</td>
                                <td>
                                    <div>新規: <strong style="font-weight: bold;">${account.new_count}</strong>件</div>
                                    <div>更新: ${account.update_count}件</div>
                                </td>
                                <td>${account.new_count + account.update_count}件</td>
                            `;
                        });
                        
                        // 支店合計行（複数アカウントがある場合のみ）
                        if (area.accounts.length > 1) {
                            const totalRow = tbody.insertRow();
                            totalRow.classList.add('area-total');
                            totalRow.innerHTML = `
                                <td colspan="3" style="padding-left: 20px; font-weight: bold; color: #2196F3;">
                                    【${area.name} 合計】
                                </td>
                                <td style="font-weight: bold; color: #2196F3;">
                                    ${area.new_count + area.update_count}件
                                </td>
                            `;
                        }
                    } else {
                        // アカウントがない支店
                        areaRow.innerHTML = `
                            <td colspan="4" style="background-color: #f5f5f5; font-weight: bold; padding: 12px; color: #666;">
                                📍 ${area.name} (ID: ${area.id})
                                <span style="float: right; color: #999;">
                                    関連アカウントなし
                                </span>
                            </td>
                        `;
                        
                        // 説明行
                        const noAccountRow = tbody.insertRow();
                        noAccountRow.innerHTML = `
                            <td colspan="4" style="padding-left: 20px; color: #999; font-style: italic;">
                                └─ この支店には関連するアカウントが設定されていません
                            </td>
                        `;
                    }
                    
                    // 区切り行
                    const separatorRow = tbody.insertRow();
                    separatorRow.innerHTML = `
                        <td colspan="4" style="height: 10px; border: none;"></td>
                    `;
                });
            }
        }
        
        function updateAccountTable(accounts) {
            const tbody = document.querySelector('#accountTable tbody');
            if (tbody) {
                tbody.innerHTML = '';
                accounts.forEach(account => {
                    const row = tbody.insertRow();
                    if (account.needs_hellowork) {
                        row.classList.add('hellowork-enabled');
                    }
                    row.innerHTML = `
                        <td>${account.id}</td>
                        <td><strong>${account.name}</strong></td>
                        <td>${account.needs_hellowork ? '<span class="status-success">✅</span>' : '<span class="status-error">❌</span>'}</td>
                        <td>${account.needs_tabelog ? '<span class="status-success">✅</span>' : '<span class="status-error">❌</span>'}</td>
                        <td>${account.needs_kanri ? '<span class="status-success">✅</span>' : '<span class="status-error">❌</span>'}</td>
                        <td>${account.area_count}支店</td>
                    `;
                });
            }
        }
        
        function updateMappingSection(mapping) {
            // マッピングセクションの更新（簡略化）
            const mappingSection = document.querySelector('#mappingSection');
            if (mappingSection && mapping.length > 0) {
                let html = '<h3>更新済み</h3><ul>';
                mapping.slice(0, 5).forEach(item => {
                    html += `<li><strong>${item.account_name}</strong> (${item.area_name})</li>`;
                });
                html += '</ul>';
                mappingSection.innerHTML = html;
            }
        }
        
        // ページ読み込み時の処理（軽量化）
        document.addEventListener('DOMContentLoaded', function() {
            // 自動読み込みを3秒後に遅延（ページ表示を高速化）
            setTimeout(loadDataWithFilter, 3000);
            
            // 今日の日付を初期設定
            const today = new Date().toISOString().split('T')[0];
            const startDateElement = document.getElementById('startDate');
            const endDateElement = document.getElementById('endDate');
            
            if (startDateElement) startDateElement.value = today;
            if (endDateElement) endDateElement.value = today;
        });

        // 日付範囲設定関数
        function setDateRange(range) {
            const today = new Date();
            const startDateInput = document.getElementById('startDate');
            const endDateInput = document.getElementById('endDate');
            
            if (range === 'today') {
                const todayStr = today.toISOString().split('T')[0];
                startDateInput.value = todayStr;
                endDateInput.value = todayStr;
            } else if (range === 'week') {
                const weekAgo = new Date(today);
                weekAgo.setDate(today.getDate() - 7);
                startDateInput.value = weekAgo.toISOString().split('T')[0];
                endDateInput.value = today.toISOString().split('T')[0];
            } else if (range === 'month') {
                const monthAgo = new Date(today);
                monthAgo.setDate(today.getDate() - 30);
                startDateInput.value = monthAgo.toISOString().split('T')[0];
                endDateInput.value = today.toISOString().split('T')[0];
            }
        }
        
        // 日付範囲でデータ表示
        async function loadDataByDateRange() {
            const startDate = document.getElementById('startDate').value;
            const endDate = document.getElementById('endDate').value;
            
            if (!startDate || !endDate) {
                alert('開始日と終了日を選択してください。');
                return;
            }
            
            const displayArea = document.getElementById('dateRangeResults');
            const titleElement = document.getElementById('dateRangeTitle');
            const contentElement = document.getElementById('dateRangeContent');
            
            displayArea.style.display = 'block';
            titleElement.textContent = 'データ読み込み中...';
            contentElement.innerHTML = '<div style="text-align: center; padding: 20px;">📊 データを取得しています...</div>';
            
            try {
                const response = await fetch('/api/date-range-data', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        start_date: startDate,
                        end_date: endDate
                    })
                });
                
                if (response.ok) {
                    const data = await response.json();
                    
                    if (data.status === 'success') {
                        titleElement.textContent = `📊 ${data.period_text}のデータ（総計: ${data.total_all}件）`;
                        
                        let html = `
                            <div style="margin-bottom: 20px; padding: 15px; background-color: #e7f3ff; border-radius: 5px;">
                                <h4>📈 期間別集計</h4>
                                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 15px; margin-top: 10px;">
                                    <div style="text-align: center; padding: 10px; background-color: #d4edda; border-radius: 5px;">
                                        <div style="font-size: 1.5em; font-weight: bold; color: #155724;">${data.total_new}</div>
                                        <div style="color: #155724; font-weight: bold;">新規データ</div>
                                    </div>
                                    <div style="text-align: center; padding: 10px; background-color: #f8f9fa; border-radius: 5px;">
                                        <div style="font-size: 1.5em; color: #856404;">${data.total_update}</div>
                                        <div style="color: #856404;">更新データ</div>
                                    </div>
                                    <div style="text-align: center; padding: 10px; background-color: #fde2e4; border-radius: 5px;">
                                        <div style="font-size: 1.5em; color: #721c24;">${data.total_unassigned}</div>
                                        <div style="color: #721c24;">振り分けなし</div>
                                    </div>
                                    <div style="text-align: center; padding: 10px; background-color: #d1ecf1; border-radius: 5px;">
                                        <div style="font-size: 1.5em; color: #0c5460;">${data.total_all}</div>
                                        <div style="color: #0c5460;">合計</div>
                                    </div>
                                </div>
                            </div>
                        `;
                        
                        // 支店別詳細
                        html += '<div style="margin-top: 20px;"><h4>🏢 支店別詳細</h4>';
                        
                        data.areas.forEach(area => {
                            html += `
                                <div style="margin-bottom: 15px; border: 1px solid #ddd; border-radius: 5px; overflow: hidden;">
                                    <div style="background-color: #f8f9fa; padding: 10px; font-weight: bold; border-bottom: 1px solid #ddd;">
                                        ${area.area_name} (新規: ${area.area_new_total}件、更新: ${area.area_update_total}件、振り分けなし: ${area.area_unassigned_total}件、合計: ${area.area_total}件)
                                    </div>
                                    <div style="padding: 10px;">
                                        <table style="width: 100%; border-collapse: collapse;">
                                            <thead>
                                                <tr style="background-color: #f0f0f0;">
                                                    <th style="padding: 8px; border: 1px solid #ddd; text-align: left;">アカウント</th>
                                                    <th style="padding: 8px; border: 1px solid #ddd; text-align: center;">新規</th>
                                                    <th style="padding: 8px; border: 1px solid #ddd; text-align: center;">更新</th>
                                                    <th style="padding: 8px; border: 1px solid #ddd; text-align: center;">合計</th>
                                                </tr>
                                            </thead>
                                            <tbody>
                            `;
                            
                            area.accounts.forEach(account => {
                                html += `
                                    <tr>
                                        <td style="padding: 8px; border: 1px solid #ddd;">${account.account_name}</td>
                                        <td style="padding: 8px; border: 1px solid #ddd; text-align: center; background-color: #d4edda;"><strong>${account.new_count}</strong></td>
                                        <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">${account.update_count}</td>
                                        <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">${account.total_count}</td>
                                    </tr>
                                `;
                            });
                            
                            html += `
                                            </tbody>
                                        </table>
                                    </div>
                                </div>
                            `;
                        });
                        
                        html += '</div>';
                        contentElement.innerHTML = html;
                        
                    } else {
                        titleElement.textContent = 'エラーが発生しました';
                        contentElement.innerHTML = `<div style="color: #dc3545; padding: 20px;">❌ ${data.message}</div>`;
                    }
                } else {
                    titleElement.textContent = 'データ取得エラー';
                    contentElement.innerHTML = '<div style="color: #dc3545; padding: 20px;">❌ サーバーエラーが発生しました</div>';
                }
            } catch (error) {
                titleElement.textContent = 'エラーが発生しました';
                contentElement.innerHTML = `<div style="color: #dc3545; padding: 20px;">❌ ${error.message}</div>`;
            }
        }
        
        // 日付範囲でExcel出力
        async function exportExcelByDateRange() {
            const startDate = document.getElementById('startDate').value;
            const endDate = document.getElementById('endDate').value;
            
            if (!startDate || !endDate) {
                alert('開始日と終了日を選択してください。');
                return;
            }
            
            try {
                const response = await fetch('/api/export-date-range', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        start_date: startDate,
                        end_date: endDate
                    })
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `hellowork_data_${startDate}_to_${endDate}.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                    
                    alert(`✅ ${startDate}〜${endDate}のデータでExcel出力が完了しました！`);
                } else {
                    alert('❌ Excel出力に失敗しました');
                }
            } catch (error) {
                alert('❌ Excel出力でエラーが発生しました: ' + error.message);
            }
        }

        // アコーディオン機能
        function toggleAccordion(sectionId) {
            const section = document.getElementById(sectionId);
            const icon = document.getElementById('branchToggleIcon');
            
            if (section.style.display === 'none') {
                section.style.display = 'block';
                icon.textContent = '▲';
            } else {
                section.style.display = 'none';
                icon.textContent = '▼';
            }
        }

        // 旧関数との互換性維持
        async function exportMapping() {
            return await exportHierarchicalReport();
        }
    </script>
</body>
</html>
'''

# ========================
# ルート定義
# ========================

@app.route('/')
def index():
    try:
        # 統計情報取得
        summary = get_area_account_summary()
        mapping = get_area_account_mapping()
        
        stats = {
            'total_areas': len(summary['areas']),
            'total_accounts': len(summary['accounts']),
            'hellowork_accounts': len([acc for acc in summary['accounts'] if acc['needs_hellowork']]),
            'total_companies': sum(area['company_count'] for area in summary['areas'])
        }
        
        return render_template_string(MAIN_TEMPLATE,
                                    stats=stats,
                                    areas=summary['areas'],
                                    accounts=summary['accounts'],
                                    area_account_mapping=mapping)
                                    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/areas')
def get_areas():
    """支店一覧取得API"""
    try:
        areas = FmArea.query.order_by(FmArea.id).all()
        return jsonify({
            'status': 'success',
            'data': [area.to_dict() for area in areas]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/accounts')
def get_accounts():
    """アカウント一覧取得API"""
    try:
        accounts = FmAccount.query.order_by(FmAccount.sort_order).all()
        return jsonify({
            'status': 'success',
            'data': [account.to_dict() for account in accounts]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/mapping')
def get_mapping():
    """支店・アカウントマッピング取得API"""
    try:
        mapping = get_area_account_mapping()
        return jsonify({
            'status': 'success',
            'data': mapping
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/export-mapping', methods=['POST'])
def export_mapping():
    """階層構造 Excel出力API（期間指定対応）"""
    try:
        # リクエストから期間指定パラメータを取得
        data = request.get_json() or {}
        date_filter = data.get('date_filter', 'today')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # 日付文字列をdateオブジェクトに変換
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # 期間指定でデータを生成
        hierarchical_data = generate_hierarchical_excel_data(
            date_filter=date_filter,
            start_date=start_date,
            end_date=end_date
        )
        
        # DataFrame作成
        df = pd.DataFrame(hierarchical_data)
        
        # Excelファイル作成
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ハローワーク送信状況')
            
            # スタイル調整
            workbook = writer.book
            worksheet = writer.sheets['ハローワーク送信状況']
            
            # フォントとスタイルの定義
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # ヘッダーのスタイル
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            # レベル別のスタイル定義
            level1_font = Font(bold=True, size=14, color='000080')  # 支店
            level1_fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
            
            level2_font = Font(bold=True, size=11, color='000000')  # アカウント
            level2_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
            
            level3_font = Font(size=10, color='333333')  # 新規/更新
            level3_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            subtotal_font = Font(bold=True, size=10, color='006600')  # 小計
            subtotal_fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
            
            # 罫線の定義
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ヘッダー行のスタイル設定
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # データ行のスタイル設定（画像フォーマット準拠）
            for row_num, row_data in enumerate(hierarchical_data, start=2):
                level = row_data.get('レベル', 0)
                item_name = row_data.get('項目名', '')
                type_name = row_data.get('種別', '')
                
                # レベルに応じてスタイルを適用
                if level == 1:  # 支店ヘッダー・支店合計
                    if '合計' in item_name:
                        # 支店合計行
                        font = Font(bold=True, size=11, color='000080')
                        fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                    else:
                        # 支店ヘッダー行
                        font = Font(bold=True, size=12, color='000080')
                        fill = PatternFill(start_color='D4E6F1', end_color='D4E6F1', fill_type='solid')
                elif level == 2:  # アカウントヘッダー
                    font = Font(bold=True, size=11, color='000000')
                    fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                elif level == 3:  # 新規/更新/小計
                    if type_name == '小計':
                        font = Font(bold=True, size=10, color='006600')
                        fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
                    elif type_name == '新規':
                        font = Font(size=10, color='0066CC')
                        fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    elif type_name == '更新':
                        font = Font(size=10, color='CC6600')
                        fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    else:
                        font = Font(size=10, color='333333')
                        fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                else:  # 区切り行など
                    font = Font(size=8)
                    fill = PatternFill()
                
                # 行の全セルにスタイルを適用
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = font
                    cell.fill = fill
                    cell.border = thin_border
                    
                    # 件数列は右寄せ、その他は左寄せ
                    if col_num == 4:  # 件数列（D列）
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 支店ヘッダー行とアカウントヘッダー行の場合、件数列を空にする
                    if (level in [1, 2] and not ('合計' in item_name)) and col_num == 4:
                        cell.value = ''
            
            # 列幅の調整
            column_widths = {
                'A': 5,   # レベル
                'B': 35,  # 項目名
                'C': 10,  # 種別
                'D': 10,  # 件数
                'E': 25   # 備考
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # フリーズペイン（ヘッダー行を固定）
            worksheet.freeze_panes = 'A2'
            
            # シートタブの色
            worksheet.sheet_properties.tabColor = "366092"
        
        output.seek(0)
        
        # ファイル名生成
        today = date.today()
        filename = f"hellowork_hierarchical_report_{today.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/filtered-data', methods=['POST'])
def get_filtered_data():
    """期間フィルタを適用したデータ取得API（軽量化対応）"""
    try:
        data = request.get_json() or {}
        date_filter = data.get('date_filter', 'today')
        
        # 全支店と関連アカウント情報を取得
        try:
            areas_with_accounts = get_all_areas_with_accounts()
        except Exception as e:
            # マッピング取得に失敗した場合は空のレスポンスを返す
            return jsonify({
                'status': 'success',
                'period': '今日',
                'period_text': '今日',
                'total_new': 0,
                'total_update': 0,
                'total_companies': 0,
                'areas': [],
                'message': f'マッピングデータの取得に失敗しました: {e}'
            })
        
        # 期間の計算（表示用）
        if date_filter == 'today':
            period_text = "今日"
        elif date_filter == 'week':
            period_text = "1週間"
        elif date_filter == 'month':
            period_text = "1ヶ月"
        elif date_filter == 'all':
            period_text = "全データ"
        else:
            period_text = "今日"
        
        # 軽量化: 支店の詳細データを構築（最初の5支店のみ）
        total_new = 0
        total_update = 0
        total_unassigned = 0
        areas_data = []
        
        # 軽量化: 処理する支店数を制限（最初の5支店のみ処理）
        limited_areas = areas_with_accounts[:5] if len(areas_with_accounts) > 5 else areas_with_accounts
        
        for area_info in limited_areas:
            area_new_total = 0
            area_update_total = 0
            area_unassigned_total = 0
            accounts_detail = []
            
            # 支店レベルでの振り分けなしデータを直接取得
            try:
                area_unassigned_total = db.session.query(Company).filter(
                    Company.fm_area_id == area_info['area_id'],
                    (Company.imported_fm_account_id.is_(None) | (Company.imported_fm_account_id == 0)),
                    Company.fm_import_result == 0,
                    func.date(Company.created_at) == datetime.now().date() if date_filter == 'today' else True
                ).count()
                
                # 期間フィルタを適用
                if date_filter == 'week':
                    start_date = datetime.now().date() - timedelta(days=7)
                    area_unassigned_total = db.session.query(Company).filter(
                        Company.fm_area_id == area_info['area_id'],
                        (Company.imported_fm_account_id.is_(None) | (Company.imported_fm_account_id == 0)),
                        Company.fm_import_result == 0,
                        func.date(Company.created_at).between(start_date, datetime.now().date())
                    ).count()
                elif date_filter == 'month':
                    start_date = datetime.now().date() - timedelta(days=30)
                    area_unassigned_total = db.session.query(Company).filter(
                        Company.fm_area_id == area_info['area_id'],
                        (Company.imported_fm_account_id.is_(None) | (Company.imported_fm_account_id == 0)),
                        Company.fm_import_result == 0,
                        func.date(Company.created_at).between(start_date, datetime.now().date())
                    ).count()
                
            except Exception as e:
                print(f"支店レベル振り分けなしデータ取得エラー: 支店{area_info['area_id']}: {e}")
                area_unassigned_total = 0
            
            # 軽量化: アカウント処理（最初の3アカウントのみ）
            for account_info in area_info['accounts'][:3]:
                try:
                    result = get_companies_data_by_period(
                        area_info['area_id'],
                        account_info['account_id'],
                        date_filter=date_filter
                    )
                    
                    account_new = result['new_count']
                    account_update = result['update_count']
                    account_unassigned = result['unassigned_count']  # 常に0になる
                    account_total = account_new + account_update  # 振り分けなしを含めない
                    
                    area_new_total += account_new
                    area_update_total += account_update
                    # area_unassigned_totalは支店レベルで直接取得するため、ここでは加算しない
                    
                    # アカウント詳細情報
                    accounts_detail.append({
                        'id': account_info['account_id'],
                        'name': account_info['account_name'],
                        'relation_type': "メイン",  # is_related=1のみ取得しているため
                        'new_count': account_new,
                        'update_count': account_update,
                        'unassigned_count': account_unassigned,
                        'total_count': account_total,  # 新規+更新のみ
                        'needs_hellowork': account_info['needs_hellowork']
                    })
                    
                except Exception as e:
                    print(f"データ取得エラー: 支店{area_info['area_id']}, アカウント{account_info['account_id']}: {e}")
                    # エラー時は0で補完
                    accounts_detail.append({
                        'id': account_info['account_id'],
                        'name': account_info['account_name'],
                        'relation_type': "メイン",
                        'new_count': 0,
                        'update_count': 0,
                        'unassigned_count': 0,
                        'total_count': 0,
                        'needs_hellowork': account_info['needs_hellowork']
                    })
                    continue
            
            total_new += area_new_total
            total_update += area_update_total
            total_unassigned += area_unassigned_total
            
            # 支店詳細情報
            areas_data.append({
                'id': area_info['area_id'],
                'name': area_info['area_name'],
                'new_count': area_new_total,
                'update_count': area_update_total,
                'unassigned_count': area_unassigned_total,
                'total_count': area_new_total + area_update_total + area_unassigned_total,  # 振り分けなしを含める
                'accounts': accounts_detail,
                'has_hellowork_accounts': area_info['has_hellowork_accounts']
            })
        
        # レスポンスデータを構築（軽量化情報付き）
        response_data = {
            'status': 'success',
            'period': period_text,
            'period_text': period_text,
            'total_new': total_new,
            'total_update': total_update,
            'total_unassigned': total_unassigned,
            'total_companies': total_new + total_update + total_unassigned,
            'areas': areas_data,
            'performance_note': f'軽量化モード: 最初の{len(limited_areas)}支店、各3アカウントまで表示'
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/date-range-data', methods=['POST'])
def get_date_range_data():
    """日付範囲指定データ取得API（支店・アカウント別）"""
    try:
        data = request.get_json() or {}
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        
        if not start_date_str or not end_date_str:
            return jsonify({'status': 'error', 'message': '開始日と終了日を指定してください'}), 400
        
        # 日付文字列をdateオブジェクトに変換
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # 支店とアカウントのマッピングを取得
        mapping = get_area_account_mapping()
        
        # 日付範囲の表示テキスト
        period_text = f"{start_date} 〜 {end_date}"
        
        # 日付範囲データを取得
        period_results = []
        total_new = 0
        total_update = 0
        total_unassigned = 0
        
        # 支店ごとにグループ化
        areas = {}
        for item in mapping:
            area_name = item['area_name']
            if area_name not in areas:
                areas[area_name] = {
                    'area_id': item['area_id'],
                    'accounts': []
                }
            areas[area_name]['accounts'].append({
                'account_id': item['account_id'],
                'account_name': item['account_name']
            })
        
        for area_name, area_data in areas.items():
            area_new_total = 0
            area_update_total = 0
            area_unassigned_total = 0
            account_details = []
            
            # 支店レベルでの振り分けなしデータを直接取得
            try:
                area_unassigned_total = db.session.query(Company).filter(
                    Company.fm_area_id == area_data["area_id"],
                    (Company.imported_fm_account_id.is_(None) | (Company.imported_fm_account_id == 0)),
                    Company.fm_import_result == 0,
                    func.date(Company.created_at).between(start_date, end_date)
                ).count()
            except Exception as e:
                print(f"支店レベル振り分けなしデータ取得エラー（日付範囲）: 支店{area_data['area_id']}: {e}")
                area_unassigned_total = 0
            
            for account in area_data['accounts']:
                # 実際のデータベースから件数を取得
                data_result = get_companies_data_by_period(
                    area_data["area_id"], 
                    account["account_id"],
                    date_filter='custom',
                    start_date=start_date,
                    end_date=end_date
                )
                
                new_count = data_result['new_count']
                update_count = data_result['update_count']
                unassigned_count = data_result['unassigned_count']  # 常に0
                
                area_new_total += new_count
                area_update_total += update_count
                # area_unassigned_totalは支店レベルで直接取得するため、ここでは加算しない
                total_new += new_count
                total_update += update_count
                
                account_details.append({
                    'account_name': account['account_name'],
                    'new_count': new_count,
                    'update_count': update_count,
                    'unassigned_count': unassigned_count,  # 常に0
                    'total_count': new_count + update_count  # 振り分けなしを含めない
                })
            
            # 支店レベルの振り分けなしを全体の合計に追加
            total_unassigned += area_unassigned_total
            
            period_results.append({
                'area_name': area_name,
                'area_new_total': area_new_total,
                'area_update_total': area_update_total,
                'area_unassigned_total': area_unassigned_total,
                'area_total': area_new_total + area_update_total + area_unassigned_total,  # 振り分けなしを含める
                'accounts': account_details
            })
        
        return jsonify({
            'status': 'success',
            'period_text': period_text,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'total_new': total_new,
            'total_update': total_update,
            'total_unassigned': total_unassigned,
            'total_all': total_new + total_update + total_unassigned,
            'areas': period_results
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/export-date-range', methods=['POST'])
def export_date_range():
    """日付範囲指定Excel出力API"""
    try:
        data = request.get_json() or {}
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        
        if not start_date_str or not end_date_str:
            return jsonify({'status': 'error', 'message': '開始日と終了日を指定してください'}), 400
        
        # 日付文字列をdateオブジェクトに変換
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # 期間指定でデータを生成
        hierarchical_data = generate_hierarchical_excel_data(
            date_filter='custom',
            start_date=start_date,
            end_date=end_date
        )
        
        # DataFrame作成
        df = pd.DataFrame(hierarchical_data)
        
        # Excelファイル作成
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ハローワーク送信状況')
            
            # スタイル調整（簡略化版）
            workbook = writer.book
            worksheet = writer.sheets['ハローワーク送信状況']
            
            # 列幅の調整
            worksheet.column_dimensions['A'].width = 5   # レベル
            worksheet.column_dimensions['B'].width = 35  # 項目名
            worksheet.column_dimensions['C'].width = 10  # 種別
            worksheet.column_dimensions['D'].width = 10  # 件数
            worksheet.column_dimensions['E'].width = 25  # 備考
        
        output.seek(0)
        
        # ファイル名生成
        filename = f"hellowork_data_{start_date_str}_to_{end_date_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/test')
def api_test():
    """API接続テスト"""
    try:
        with db.engine.connect() as connection:
            result = connection.execute(text('SELECT VERSION()'))
            mysql_version = result.fetchone()[0]
        
        # 実データ統計
        stats = get_area_account_summary()
        
        return jsonify({
            'status': 'success',
            'message': 'API is working with real data',
            'database': 'scraping',
            'mysql_version': mysql_version,
            'environment': os.getenv('FLASK_ENV', 'production'),
            'data_summary': {
                'areas': len(stats['areas']),
                'accounts': len(stats['accounts']),
                'total_companies': sum(area['company_count'] for area in stats['areas'])
            }
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'database': 'disconnected'
        }), 500

@app.route('/api/debug-unassigned', methods=['GET'])
def debug_unassigned():
    """振り分けなしデータのデバッグ用エンドポイント"""
    try:
        # 10月15日のfm_import_result = 0のデータを詳細確認
        today_unassigned = db.session.query(
            Company.fm_area_id,
            Company.imported_fm_account_id,
            func.count(Company.id).label('count')
        ).filter(
            Company.fm_import_result == 0,
            func.date(Company.created_at) == '2025-10-15'
        ).group_by(
            Company.fm_area_id,
            Company.imported_fm_account_id
        ).order_by(
            Company.fm_area_id,
            Company.imported_fm_account_id
        ).all()
        
        # 各支店の最初のアカウントを確認
        area_first_accounts = {}
        area_account_mapping = db.session.query(
            FmAreaAccount.fm_area_id,
            FmAreaAccount.fm_account_id,
            FmAccount.department_name
        ).join(
            FmAccount, FmAreaAccount.fm_account_id == FmAccount.id
        ).filter(
            FmAreaAccount.is_related == 1
        ).order_by(
            FmAreaAccount.fm_area_id,
            FmAccount.sort_order
        ).all()
        
        for mapping in area_account_mapping:
            area_id = mapping.fm_area_id
            if area_id not in area_first_accounts:
                area_first_accounts[area_id] = {
                    'account_id': mapping.fm_account_id,
                    'account_name': mapping.department_name
                }
        
        return jsonify({
            'status': 'success',
            'today_unassigned_by_area_account': [
                {
                    'fm_area_id': row.fm_area_id,
                    'imported_fm_account_id': row.imported_fm_account_id,
                    'count': row.count
                } for row in today_unassigned
            ],
            'area_first_accounts': area_first_accounts
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)